import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def add_data_to_excel(data, title):
    workbook = load_workbook(rf'.\docs\{title}.xlsx')
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(rf'.\docs\{title}.xlsx')


def create_excel(category):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Название'
    ws['B1'] = 'Автор'
    ws['C1'] = 'Цена'
    ws['D1'] = 'Просмотры'
    ws['E1'] = 'Лайки'
    ws['F1'] = 'Кол-во наград'
    ws['G1'] = 'Жанр 1'
    ws['H1'] = 'Жанр 2'
    ws['I1'] = 'Жанр 3'
    ws['J1'] = 'Тег 1'
    ws['K1'] = 'Тег 2'
    ws['L1'] = 'Тег 3'
    ws['M1'] = 'Тег 4'
    ws['N1'] = 'Тег 5'
    ws['O1'] = 'Тег 6'
    ws['P1'] = 'Тег 7'
    ws['Q1'] = 'Тег 8'

    wb.save(rf'.\docs\{category}.xlsx')
